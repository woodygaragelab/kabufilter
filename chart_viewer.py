import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.patches import Rectangle
from matplotlib.lines import Line2D
import os
import signal
import threading
import yfinance as yf
import matplotlib

matplotlib.rcParams["font.family"] = ["Yu Gothic", "Meiryo", "MS Gothic", "DejaVu Sans"]

CHART_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(CHART_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)


def load_csv(code: str) -> pd.DataFrame:
    filename = os.path.join(DATA_DIR, f"{code}.csv")
    if not os.path.exists(filename):
        raise FileNotFoundError(f"ファイルが見つかりません: {filename}")

    # yfinance出力CSVは先頭にPrice/Ticker/Dateの3行ヘッダがあるため2・3行目をスキップ
    df = pd.read_csv(filename, skiprows=[1, 2], index_col=0)
    df.index.name = "Date"
    df = df.reset_index()
    df["Date"] = pd.to_datetime(df["Date"])
    df = df.dropna(subset=["Date", "Open", "High", "Low", "Close"])
    if df.empty:
        raise ValueError("有効なデータがありません")

    df = df.sort_values("Date").reset_index(drop=True)
    return df


def calc_stats(df: pd.DataFrame) -> dict:
    """終値・移動平均・売買高・トレンド判定を計算して辞書で返す。"""
    def _r(v):
        return round(float(v), 1) if pd.notna(v) else None
    def _b(v):
        return bool(v) if pd.notna(v) else None

    close = df["Close"].iloc[-1]
    ma25  = df["Close"].rolling(25).mean().iloc[-1]
    ma75  = df["Close"].rolling(75).mean().iloc[-1]
    tv    = df["Volume"] * df["Close"] / 1e8
    tv25  = tv.rolling(25).mean().iloc[-1]
    tv75  = tv.rolling(75).mean().iloc[-1]

    close_above_ma25 = _b(close > ma25)          if pd.notna(ma25) else None
    ma25_above_ma75  = _b(ma25  > ma75)           if pd.notna(ma75) else None
    trend            = _b(close > ma25 > ma75)    if pd.notna(ma25) and pd.notna(ma75) else None
    tv75_above_10    = _b(tv75 > 10)              if pd.notna(tv75) else None

    return {
        "Close":          _r(close),
        "MA25":           _r(ma25),
        "MA75":           _r(ma75),
        "Turnover":       _r(tv.iloc[-1]),
        "TurnoverMA25":   _r(tv25),
        "TurnoverMA75":   _r(tv75),
        "CloseAboveMA25": close_above_ma25,
        "MA25AboveMA75":  ma25_above_ma75,
        "Trend":          trend,
        "TV75Above10":    tv75_above_10,
    }


def draw_candlestick(ax_price, ax_vol, df: pd.DataFrame, code: str, name: str = "", ratio: float = 2.0):
    ax_price.clear()
    ax_vol.clear()

    dates = mdates.date2num(df["Date"].to_numpy().astype("datetime64[ms]").astype(object))
    width = 0.6

    for date, row in zip(dates, df.itertuples()):
        color = "#e74c3c" if row.Close >= row.Open else "#2ecc71"

        # ローソク足の実体
        body_low = min(row.Open, row.Close)
        body_high = max(row.Open, row.Close)
        rect = Rectangle(
            (date - width / 2, body_low),
            width,
            body_high - body_low,
            color=color,
            zorder=3,
        )
        ax_price.add_patch(rect)

        # ヒゲ
        ax_price.plot(
            [date, date], [row.Low, body_low], color=color, linewidth=1, zorder=2
        )
        ax_price.plot(
            [date, date], [body_high, row.High], color=color, linewidth=1, zorder=2
        )

        # 売買高（Volume × Close、億円単位）
        vol_color = "#e74c3c" if row.Close >= row.Open else "#2ecc71"
        turnover = (
            row.Volume * row.Close / 1e8
            if hasattr(row, "Volume") and pd.notna(row.Volume)
            else 0
        )
        bar = Rectangle(
            (date - width / 2, 0),
            width,
            turnover,
            color=vol_color,
            alpha=0.7,
        )
        ax_vol.add_patch(bar)

    # 移動平均線
    for window, color in [(25, "#f9e2af"), (75, "#89dceb")]:
        ma = df["Close"].rolling(window).mean()
        valid = ma.notna()
        if valid.any():
            ax_price.plot(
                dates[valid.values],
                ma[valid].values,
                color=color,
                linewidth=1.5,
                label=f"MA{window}",
                zorder=4,
            )
    ax_price.legend(fontsize=12, loc="upper left")

    # 軸設定
    ax_price.set_xlim(dates[0] - 1, dates[-1] + 1)
    price_max = df["High"].max()
    margin = price_max * 0.02
    y_top = price_max + margin
    y_bottom = (df["Low"].min() - margin) if ratio <= 1.0 else (y_top / ratio)
    ax_price.set_ylim(y_bottom, y_top)
    ax_price.set_title(f"{code}  {name}", fontsize=19)
    ax_price.set_ylabel("価格", fontsize=16)
    ax_price.tick_params(labelsize=15)
    ax_price.xaxis.set_major_formatter(mdates.DateFormatter("%y/%m/%d"))
    ax_price.xaxis.set_major_locator(mdates.AutoDateLocator())
    ax_price.grid(True, alpha=0.3)

    ax_vol.set_xlim(ax_price.get_xlim())
    if "Volume" in df.columns and df["Volume"].notna().any():
        turnover = df["Volume"] * df["Close"] / 1e8
        turnover_max = turnover.max()
        ax_vol.set_ylim(0, turnover_max * 1.2)
        for window, color in [(25, "#f9e2af"), (75, "#89dceb")]:
            ma = turnover.rolling(window).mean()
            valid = ma.notna()
            if valid.any():
                ax_vol.plot(
                    dates[valid.values],
                    ma[valid].values,
                    color=color,
                    linewidth=1.5,
                    label=f"MA{window}",
                    zorder=4,
                )
        ax_vol.legend(fontsize=12, loc="upper left")
    ax_vol.set_ylabel("売買高(億円)", fontsize=16)
    ax_vol.tick_params(labelsize=15)
    ax_vol.xaxis.set_major_formatter(mdates.DateFormatter("%y/%m/%d"))
    ax_vol.xaxis.set_major_locator(mdates.AutoDateLocator())
    ax_vol.grid(True, alpha=0.3)

    plt.setp(ax_price.get_xticklabels(), rotation=30, ha="right")
    plt.setp(ax_vol.get_xticklabels(), rotation=30, ha="right")


class ChartApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("株価チャートビューア")
        self.geometry("1400x800")
        self.configure(bg="#1e1e2e")
        self._xlsx_lock            = threading.Lock()
        self._codes                = []
        self._filtered_codes       = []
        self._programmatic_select  = False
        self._build_ui()

    def _build_ui(self):
        # 上部入力エリア
        top = tk.Frame(self, bg="#1e1e2e", pady=8)
        top.pack(fill=tk.X, padx=12)

        self.code_var = tk.StringVar(value="7203")

        tk.Label(top, text="倍率:", bg="#1e1e2e", fg="#cdd6f4", font=("", 17)).pack(
            side=tk.LEFT, padx=(16, 0)
        )
        self.ratio_var = tk.StringVar(value="2.0")
        ratio_cb = ttk.Combobox(
            top, textvariable=self.ratio_var,
            values=["1.5", "2.0", "3.0", "4.0"], width=4, font=("", 17), state="readonly"
        )
        ratio_cb.pack(side=tk.LEFT, padx=4)
        ratio_cb.bind("<<ComboboxSelected>>", lambda _: self._show_chart())

        tk.Label(top, text="分類:", bg="#1e1e2e", fg="#cdd6f4", font=("", 17)).pack(
            side=tk.LEFT, padx=(16, 0)
        )
        self.category_var = tk.StringVar()
        cat_entry = ttk.Entry(top, textvariable=self.category_var, width=12, font=("", 17))
        cat_entry.pack(side=tk.LEFT, padx=4)
        cat_entry.bind("<Return>", lambda _: self._save_category())

        cat_btn = tk.Button(
            top,
            text="保存",
            command=self._save_category,
            bg="#a6e3a1",
            fg="#1e1e2e",
            font=("", 17, "bold"),
            relief=tk.FLAT,
            padx=10,
        )
        cat_btn.pack(side=tk.LEFT)

        self.status = tk.Label(
            top, text="", bg="#1e1e2e", fg="#f38ba8", font=("", 16)
        )
        self.status.pack(side=tk.LEFT, padx=10)

        # メインエリア（左リスト＋右チャート）
        content = tk.Frame(self, bg="#1e1e2e")
        content.pack(fill=tk.BOTH, expand=True)

        # 左パネル：銘柄リスト
        left = tk.Frame(content, bg="#181825", width=370)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(12, 0), pady=4)
        left.pack_propagate(False)

        tk.Label(
            left, text="銘柄リスト", bg="#181825", fg="#cdd6f4", font=("", 16, "bold")
        ).pack(pady=(6, 2))

        # 列ごとのフィルタ入力欄
        self._filter_frame = tk.Frame(left, bg="#181825")
        self._filter_frame.pack(fill=tk.X, padx=4, pady=(0, 0))

        self.code_filter_var = tk.StringVar()
        self.name_filter_var = tk.StringVar()
        self.cat_filter_var  = tk.StringVar()

        ttk.Entry(self._filter_frame, textvariable=self.code_filter_var, font=("", 13), width=1).grid(
            row=0, column=0, sticky="ew")
        ttk.Entry(self._filter_frame, textvariable=self.name_filter_var, font=("", 13), width=1).grid(
            row=0, column=1, sticky="ew")
        ttk.Entry(self._filter_frame, textvariable=self.cat_filter_var,  font=("", 13), width=1).grid(
            row=0, column=2, sticky="ew")
        # スクロールバー幅分のスペーサー（後で実幅に更新）
        self._sb_spacer = tk.Frame(self._filter_frame, bg="#181825", width=17)
        self._sb_spacer.grid(row=0, column=3)

        self._filter_frame.columnconfigure(0, minsize=62,  weight=0)
        self._filter_frame.columnconfigure(1, minsize=80,  weight=1)
        self._filter_frame.columnconfigure(2, minsize=88,  weight=0)
        self._filter_frame.columnconfigure(3, minsize=17,  weight=0)

        for var in (self.code_filter_var, self.name_filter_var, self.cat_filter_var):
            var.trace_add("write", lambda *_: self._apply_filter())

        # Treeview スタイル
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Stock.Treeview",
                        background="#1e1e2e", foreground="#cdd6f4",
                        fieldbackground="#1e1e2e", font=("", 15), rowheight=30)
        style.configure("Stock.Treeview.Heading",
                        background="#252535", foreground="#cdd6f4", font=("", 12, "bold"))
        style.map("Stock.Treeview",
                  background=[("selected", "#89b4fa")],
                  foreground=[("selected", "#1e1e2e")])

        tree_frame = tk.Frame(left, bg="#181825")
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=(0, 4))

        self._tree_sb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL)
        self._tree_sb.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree = ttk.Treeview(
            tree_frame,
            columns=("code", "name", "category"),
            show="headings",
            yscrollcommand=self._tree_sb.set,
            style="Stock.Treeview",
            selectmode="browse",
        )
        self.tree.heading("code",     text="コード")
        self.tree.heading("name",     text="銘柄名")
        self.tree.heading("category", text="分類")
        self.tree.column("code",     width=62,  minwidth=50, stretch=False)
        self.tree.column("name",     width=170, minwidth=80, stretch=True)
        self.tree.column("category", width=88,  minwidth=60, stretch=False)
        self.tree.tag_configure("today", foreground="#cdd6f4")
        self.tree.tag_configure("old",   foreground="#6c7086")
        self.tree.pack(fill=tk.BOTH, expand=True)
        self._tree_sb.config(command=self.tree.yview)

        self._codes = self._load_my_codes()
        self._filtered_codes = self._codes[:]
        self._apply_filter()
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        self.tree.bind("<Configure>", lambda _: self._sync_filter_columns())
        self.after_idle(self._sync_filter_columns)
        self.bind("<Up>",   self._on_key_nav)
        self.bind("<Down>", self._on_key_nav)

        # 右パネル：チャート
        right = tk.Frame(content, bg="#1e1e2e")
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=8, pady=4)

        plt.style.use("dark_background")
        self.fig, (self.ax_price, self.ax_vol) = plt.subplots(
            2, 1, figsize=(10, 6), gridspec_kw={"height_ratios": [3, 1]}, facecolor="#1e1e2e"
        )
        self.fig.subplots_adjust(hspace=0.05, bottom=0.12)

        self.canvas = FigureCanvasTkAgg(self.fig, master=right)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        self._ch_vp = self._ch_vv = self._ch_hp = self._ch_hv = None
        self._drag_press_x       = None
        self._drag_xlim          = None
        self._programmatic_select = False
        self.canvas.mpl_connect("button_press_event",   self._on_button_press)
        self.canvas.mpl_connect("button_release_event", self._on_button_release)
        self.canvas.mpl_connect("motion_notify_event",  self._on_mouse_move)
        self.canvas.mpl_connect("axes_leave_event",     lambda _: self._hide_crosshair())
        self.canvas.mpl_connect("scroll_event",         self._on_scroll)

        self._show_chart()

    def _read_excel(self) -> pd.DataFrame | None:
        """my_codes.xlsx を読み込む。存在しない場合は CSV から生成して再試行。"""
        path     = os.path.join(CHART_DIR, "my_codes.xlsx")
        csv_path = os.path.join(CHART_DIR, "my_codes.csv")
        with self._xlsx_lock:
            if not os.path.exists(path):
                if os.path.exists(csv_path):
                    pd.read_csv(csv_path, dtype={"Code": str}).to_excel(path, index=False)
                else:
                    return None
            try:
                return pd.read_excel(path, dtype={"Code": str, "Category": str})
            except Exception:
                return None

    def _load_my_codes(self) -> list:
        df = self._read_excel()
        if df is None:
            return []
        df["Code"] = df["Code"].str.zfill(4)
        if "Category" not in df.columns:
            df["Category"] = ""
        df["Category"] = df["Category"].fillna("")
        if "DataDate" not in df.columns:
            df["DataDate"] = ""
        df["DataDate"] = pd.to_datetime(df["DataDate"], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
        return list(zip(df["Code"], df["Name"], df["Category"], df["DataDate"]))

    def _sync_filter_columns(self):
        self.update_idletasks()
        code_w = self.tree.column("code",     "width")
        cat_w  = self.tree.column("category", "width")
        sb_w   = max(self._tree_sb.winfo_width(), 1)
        self._filter_frame.columnconfigure(0, minsize=code_w, weight=0)
        self._filter_frame.columnconfigure(2, minsize=cat_w,  weight=0)
        self._filter_frame.columnconfigure(3, minsize=sb_w,   weight=0)
        self._sb_spacer.configure(width=sb_w)

    def _on_key_nav(self, event):
        if isinstance(self.focus_get(), (ttk.Entry, tk.Entry)):
            return
        children = self.tree.get_children()
        if not children:
            return
        sel = self.tree.selection()
        if not sel:
            new_item = children[0] if event.keysym == "Down" else children[-1]
        else:
            idx = list(children).index(sel[0])
            if event.keysym == "Down":
                idx = min(idx + 1, len(children) - 1)
            else:
                idx = max(idx - 1, 0)
            new_item = children[idx]
        self._programmatic_select = True
        self.tree.selection_set(new_item)
        self.tree.see(new_item)
        self._programmatic_select = False
        code = str(self.tree.item(new_item)["values"][0]).zfill(4)
        self.code_var.set(code)
        self._load_category(code)
        self._download_and_show()
        return "break"

    def _on_tree_select(self, *_):
        if self._programmatic_select:
            return
        selection = self.tree.selection()
        if not selection:
            return
        code = str(self.tree.item(selection[0])["values"][0]).zfill(4)
        self.code_var.set(code)
        self._load_category(code)
        self._download_and_show()

    def _apply_filter(self):
        today = pd.Timestamp.today().strftime("%Y-%m-%d")
        code_q = self.code_filter_var.get().strip().lower()
        name_q = self.name_filter_var.get().strip().lower()
        cat_q  = self.cat_filter_var.get().strip().lower()
        self._filtered_codes = [
            (c, n, cat, dd) for c, n, cat, dd in self._codes
            if (not code_q or code_q in c.lower())
            and (not name_q or name_q in n.lower())
            and (not cat_q  or cat_q  in cat.lower())
        ]
        # 選択コードを保存（tree 選択 → code_var の順で参照）
        sel = self.tree.selection()
        if sel:
            selected_code = str(self.tree.item(sel[0])["values"][0]).zfill(4)
        else:
            selected_code = self.code_var.get().strip() or None

        self._programmatic_select = True
        self.tree.delete(*self.tree.get_children())
        for code, name, category, datadate in self._filtered_codes:
            tag = "today" if datadate == today else "old"
            iid = self.tree.insert("", tk.END, values=(code, name, category), tags=(tag,))
            if code == selected_code:
                self.tree.selection_set(iid)
                self.tree.see(iid)
        self._programmatic_select = False

    def _load_category(self, code: str):
        mc = self._read_excel()
        if mc is None:
            self.category_var.set("")
            return
        mc["Code"] = mc["Code"].str.zfill(4)
        row = mc[mc["Code"] == code]
        if not row.empty and "Category" in mc.columns:
            val = row["Category"].iloc[0]
            self.category_var.set("" if pd.isna(val) else str(val))
        else:
            self.category_var.set("")

    def _update_excel_row(self, code: str, updates: dict):
        import openpyxl
        path = os.path.join(CHART_DIR, "my_codes.xlsx")
        if not os.path.exists(path):
            return

        with self._xlsx_lock:
            try:
                # pandas で行番号を確実に特定
                df = pd.read_excel(path, dtype={"Code": str})
                df["Code"] = df["Code"].str.strip().str.zfill(4)
                matches = df.index[df["Code"] == code].tolist()
                if not matches:
                    return
                excel_row = matches[0] + 2  # DataFrame は 0 始まり、Excel は 1 始まり＋ヘッダ行

                wb = openpyxl.load_workbook(path)
                ws = wb.active

                # ヘッダ行から列名→列番号のマップを構築
                headers = {}
                max_col = 0
                for cell in ws[1]:
                    if cell.value is not None:
                        headers[str(cell.value).strip()] = cell.column
                        max_col = max(max_col, cell.column)

                # 存在しない列はヘッダ行末尾に追加
                for col_name in updates:
                    if col_name not in headers:
                        max_col += 1
                        ws.cell(row=1, column=max_col).value = col_name
                        headers[col_name] = max_col

                for col_name, value in updates.items():
                    ws.cell(row=excel_row, column=headers[col_name]).value = value

                wb.save(path)
            except Exception:
                pass

    def _save_category(self):
        code = self.code_var.get().strip()
        if not code:
            return
        self._update_excel_row(code, {"Category": self.category_var.get().strip()})
        self._refresh_listbox()
        self.status.config(text=f"{code} の分類を保存しました")

    def _refresh_listbox(self):
        self._codes = self._load_my_codes()
        self._apply_filter()

    def _download_and_show(self):
        code = self.code_var.get().strip()
        if not code:
            return
        today = pd.Timestamp.today().strftime("%Y-%m-%d")
        datadate = next((dd for c, *_, dd in self._codes if c == code), "")
        if datadate == today:
            self._show_chart()
            return
        self.status.config(text=f"{code} をダウンロード中...")
        self.update_idletasks()

        def task():
            try:
                ticker = f"{code}.T"
                data = yf.download(ticker, period="400d", interval="1d", progress=False)
                if data.empty:
                    self.after(0, lambda: self.status.config(text=f"データ取得失敗: {ticker}"))
                    return
                data = data.tail(275)
                data.to_csv(os.path.join(DATA_DIR, f"{code}.csv"))
                from datetime import datetime
                self._update_excel_row(code, {"DataDate": datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)})
                self.after(0, self._refresh_listbox)
                self.after(0, self._show_chart)
            except Exception as e:
                msg = f"ダウンロードエラー: {e}"
                self.after(0, lambda m=msg: self.status.config(text=m))

        threading.Thread(target=task, daemon=True).start()

    def _init_crosshair(self):
        kw = dict(color="#6c7086", linewidth=0.8, linestyle="--", visible=False, zorder=10)
        self._ch_vp = self.ax_price.axvline(0, **kw)
        self._ch_vv = self.ax_vol.axvline(0, **kw)
        self._ch_hp = self.ax_price.axhline(0, **kw)
        self._ch_hv = self.ax_vol.axhline(0, **kw)

    def _on_scroll(self, event):
        if event.inaxes not in (self.ax_price, self.ax_vol):
            return
        ratios = ["1.5", "2.0", "3.0", "4.0"]
        cur = self.ratio_var.get()
        idx = ratios.index(cur) if cur in ratios else 1
        if event.step > 0:
            idx = max(idx - 1, 0)
        else:
            idx = min(idx + 1, len(ratios) - 1)
        self.ratio_var.set(ratios[idx])
        self._show_chart()

    def _on_button_press(self, event):
        if event.button == 1 and event.inaxes in (self.ax_price, self.ax_vol):
            self._drag_press_x = event.x
            self._drag_xlim    = self.ax_price.get_xlim()
            self.canvas.get_tk_widget().config(cursor="fleur")

    def _on_button_release(self, event):
        if event.button == 1:
            self._drag_press_x = None
            self._drag_xlim    = None
            self.canvas.get_tk_widget().config(cursor="")

    def _on_mouse_move(self, event):
        # ドラッグスクロール
        if self._drag_press_x is not None and event.x is not None:
            xlim = self._drag_xlim
            ax_w = self.ax_price.get_position().width * self.fig.get_size_inches()[0] * self.fig.dpi
            if ax_w > 0:
                dx = (event.x - self._drag_press_x) * (xlim[1] - xlim[0]) / ax_w
                self.ax_price.set_xlim(xlim[0] - dx, xlim[1] - dx)
                self.ax_vol.set_xlim(xlim[0] - dx, xlim[1] - dx)
            self.canvas.draw_idle()
            return

        # クロスヘア更新
        if self._ch_vp is None:
            return
        if event.inaxes not in (self.ax_price, self.ax_vol):
            self._hide_crosshair()
            return
        x, y = event.xdata, event.ydata
        if x is None or y is None:
            return
        self._ch_vp.set_xdata([x])
        self._ch_vv.set_xdata([x])
        self._ch_vp.set_visible(True)
        self._ch_vv.set_visible(True)
        if event.inaxes == self.ax_price:
            self._ch_hp.set_ydata([y])
            self._ch_hp.set_visible(True)
            self._ch_hv.set_visible(False)
        else:
            self._ch_hv.set_ydata([y])
            self._ch_hv.set_visible(True)
            self._ch_hp.set_visible(False)
        self.canvas.draw_idle()

    def _hide_crosshair(self):
        if self._ch_vp is None:
            return
        for line in (self._ch_vp, self._ch_vv, self._ch_hp, self._ch_hv):
            line.set_visible(False)
        self.canvas.draw_idle()

    def _show_chart(self):
        code = self.code_var.get().strip()
        if not code:
            return
        try:
            df = load_csv(code)
            name = next((n for c, n, *_ in self._codes if c == code), "")
            draw_candlestick(self.ax_price, self.ax_vol, df, code, name, ratio=float(self.ratio_var.get()))
            self._init_crosshair()
            self.canvas.draw()
            self._save_stats(code, df)
            self.status.config(text=f"{len(df)}件のデータを読み込みました")
        except FileNotFoundError as e:
            self.status.config(text=str(e))
        except Exception as e:
            self.status.config(text=f"エラー: {e}")

    def _save_stats(self, code: str, df: pd.DataFrame):
        stats = calc_stats(df)
        category = next((tup[2] for tup in self._codes if tup[0] == code), "")
        add_chars = ""
        if stats.get("TV75Above10") and "1" not in category:
            add_chars += "1"
        if stats.get("Trend") and "2" not in category:
            add_chars += "2"
        if add_chars:
            stats["Category"] = "".join(sorted(set(category.strip() + add_chars)))
            self.after(0, self._refresh_listbox)
        self._update_excel_row(code, stats)


if __name__ == "__main__":
    signal.signal(signal.SIGINT, signal.SIG_DFL)
    app = ChartApp()
    app.mainloop()
